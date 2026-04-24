"""
AutoParts Express - Modelo de Reportes
HU-09: Reportes de Ventas e Inventario (Sprint 3)
HU-10: Reporte de Ventas por Vendedor   (Sprint 3)
"""

from datetime import date as date_type, datetime
from typing import Optional
from db.connection import db_cursor

# ── Constantes (evitar literales duplicados) ──────────────────────────────────
TITULO_AUTOPARTS  = "AutoParts Express"
TITULO_VENTAS     = "Total Ventas"
TITULO_INGRESOS   = "Ingresos Totales"
TITULO_TICKET     = "Ticket Promedio"
TITULO_CODIGO     = "Código"


# ── Reporte de Ventas por Período (HU-09) ─────────────────────────────────────

def reporte_ventas(fecha_inicio: date_type, fecha_fin: date_type) -> dict:
    """Genera un dict con todos los datos para el reporte de ventas."""
    with db_cursor() as cur:
        cur.execute("""
            SELECT
                COUNT(*)                    AS total_ventas,
                COALESCE(SUM(total), 0)     AS ingresos_totales,
                COALESCE(SUM(descuento), 0) AS descuentos_totales,
                COALESCE(AVG(total), 0)     AS ticket_promedio
            FROM ventas
            WHERE estado = 'Completada'
              AND fecha_hora::date BETWEEN %s AND %s
        """, (fecha_inicio, fecha_fin))
        resumen = dict(cur.fetchone())

        cur.execute("""
            SELECT fecha_hora::date AS dia, COUNT(*) AS cantidad, SUM(total) AS ingreso
            FROM ventas
            WHERE estado = 'Completada'
              AND fecha_hora::date BETWEEN %s AND %s
            GROUP BY dia ORDER BY dia ASC
        """, (fecha_inicio, fecha_fin))
        por_dia = [dict(r) for r in cur.fetchall()]

        cur.execute("""
            SELECT p.codigo, p.nombre AS producto_nombre, c.nombre AS categoria,
                   SUM(vd.cantidad) AS unidades_vendidas, SUM(vd.subtotal) AS ingresos
            FROM venta_detalles vd
            JOIN ventas    v ON v.id = vd.venta_id
            JOIN productos p ON p.id = vd.producto_id
            LEFT JOIN categorias c ON c.id = p.categoria_id
            WHERE v.estado = 'Completada'
              AND v.fecha_hora::date BETWEEN %s AND %s
            GROUP BY p.codigo, p.nombre, c.nombre
            ORDER BY unidades_vendidas DESC LIMIT 10
        """, (fecha_inicio, fecha_fin))
        productos_top = [dict(r) for r in cur.fetchall()]

        cur.execute("""
            SELECT metodo_pago, COUNT(*) AS cantidad, SUM(total) AS total
            FROM ventas
            WHERE estado = 'Completada'
              AND fecha_hora::date BETWEEN %s AND %s
            GROUP BY metodo_pago ORDER BY total DESC
        """, (fecha_inicio, fecha_fin))
        por_metodo = [dict(r) for r in cur.fetchall()]

    return {
        "periodo":         {"inicio": fecha_inicio, "fin": fecha_fin},
        "resumen":         resumen,
        "por_dia":         por_dia,
        "productos_top":   productos_top,
        "por_metodo_pago": por_metodo,
    }


# ── Reporte de Inventario Actual (HU-09) ──────────────────────────────────────

def reporte_inventario() -> dict:
    """Genera datos para el reporte de inventario actual."""
    with db_cursor() as cur:
        cur.execute("""
            SELECT COUNT(*) AS total_productos,
                   SUM(stock_actual) AS total_unidades,
                   SUM(precio_unitario * stock_actual) AS valor_inventario,
                   COUNT(*) FILTER (WHERE stock_actual <= stock_minimo) AS productos_criticos
            FROM productos WHERE activo = TRUE
        """)
        resumen = dict(cur.fetchone())

        cur.execute("""
            SELECT p.codigo, p.nombre, c.nombre AS categoria,
                   p.precio_unitario, p.stock_actual, p.stock_minimo,
                   CASE
                       WHEN p.stock_actual = 0               THEN 'Sin stock'
                       WHEN p.stock_actual <= p.stock_minimo  THEN 'Crítico'
                       WHEN p.stock_actual <= p.stock_minimo * 2 THEN 'Bajo'
                       ELSE 'Normal'
                   END AS estado_stock
            FROM productos p LEFT JOIN categorias c ON c.id = p.categoria_id
            WHERE p.activo = TRUE ORDER BY p.stock_actual ASC
        """)
        detalle = [dict(r) for r in cur.fetchall()]
        criticos = [p for p in detalle if p["estado_stock"] in ("Sin stock", "Crítico")]

    return {
        "generado_en": datetime.now(),
        "resumen":     resumen,
        "detalle":     detalle,
        "criticos":    criticos,
    }


# ── Reporte de Ventas por Vendedor (HU-10) ────────────────────────────────────

def reporte_por_vendedor(fecha_inicio: date_type, fecha_fin: date_type,
                          vendedor_id: Optional[int] = None) -> dict:
    """Genera el reporte de desempeño por vendedor."""
    params_v = [fecha_inicio, fecha_fin]
    filtro_v = ""
    if vendedor_id:
        filtro_v = "AND u.id = %s"
        params_v.append(vendedor_id)

    with db_cursor() as cur:
        cur.execute(f"""
            SELECT u.id AS vendedor_id, u.nombre AS vendedor_nombre,
                   COUNT(v.id) AS total_ventas, SUM(v.total) AS ingresos_totales,
                   AVG(v.total) AS ticket_promedio, MAX(v.total) AS venta_maxima
            FROM ventas v JOIN usuarios u ON u.id = v.vendedor_id
            WHERE v.estado = 'Completada'
              AND v.fecha_hora::date BETWEEN %s AND %s {filtro_v}
            GROUP BY u.id, u.nombre ORDER BY ingresos_totales DESC
        """, params_v)
        por_vendedor = [dict(r) for r in cur.fetchall()]

        params_d = [fecha_inicio, fecha_fin]
        filtro_d = ""
        if vendedor_id:
            filtro_d = "AND v.vendedor_id = %s"
            params_d.append(vendedor_id)

        cur.execute(f"""
            SELECT v.numero_factura, v.fecha_hora, u.nombre AS vendedor_nombre,
                   c.nombre AS cliente_nombre, v.total, v.metodo_pago
            FROM ventas v
            JOIN usuarios u ON u.id = v.vendedor_id
            JOIN clientes c ON c.id = v.cliente_id
            WHERE v.estado = 'Completada'
              AND v.fecha_hora::date BETWEEN %s AND %s {filtro_d}
            ORDER BY v.fecha_hora DESC
        """, params_d)
        detalle = [dict(r) for r in cur.fetchall()]

    return {
        "periodo":      {"inicio": fecha_inicio, "fin": fecha_fin},
        "por_vendedor": por_vendedor,
        "detalle":      detalle,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _excel_estilos(openpyxl):
    """Retorna los estilos comunes para encabezados y celdas de Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font  = Font(bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="1A2942")
    center_align = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin")
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center_align, thin_border


def _excel_escribir_header(ws, fila, columnas, header_font, header_fill,
                            center_align, thin_border):
    """Escribe una fila de encabezado con estilo en la hoja."""
    for col_idx, texto in enumerate(columnas, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=texto)
        cell.font, cell.fill = header_font, header_fill
        cell.alignment, cell.border = center_align, thin_border


def _excel_escribir_datos(ws, fila, valores, thin_border):
    """Escribe una fila de datos con bordes."""
    from openpyxl.styles import Alignment
    for col_idx, val in enumerate(valores, start=1):
        cell = ws.cell(row=fila, column=col_idx, value=val)
        cell.border    = thin_border
        cell.alignment = Alignment(vertical="center")


def _excel_hoja_ventas(ws, datos, header_font, header_fill, center_align, thin_border):
    """Rellena la hoja Excel para el reporte de ventas."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.title = "Ventas por Período"
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"Reporte de Ventas — {datos['periodo']['inicio']} al {datos['periodo']['fin']}"
    c.font      = Font(bold=True, size=14, color="1A2942")
    c.alignment = center_align
    c.fill      = PatternFill("solid", fgColor="EDF1F7")

    r = datos["resumen"]
    ws.cell(row=3, column=1, value="RESUMEN").font = Font(bold=True, size=11)
    for row_num, (label, value) in enumerate([
        (TITULO_VENTAS,    int(r["total_ventas"])),
        (TITULO_INGRESOS,  float(r["ingresos_totales"])),
        ("Descuentos",     float(r["descuentos_totales"])),
        (TITULO_TICKET,    round(float(r["ticket_promedio"]), 2)),
    ], start=4):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=value)

    ws.cell(row=9, column=1, value="TOP 10 PRODUCTOS").font = Font(bold=True, size=11)
    _excel_escribir_header(ws, 10, [TITULO_CODIGO, "Producto", "Categoría", "Unidades", "Ingresos"],
                           header_font, header_fill, center_align, thin_border)
    for i, p in enumerate(datos["productos_top"], start=11):
        _excel_escribir_datos(ws, i, [
            p["codigo"], p["producto_nombre"], p["categoria"],
            int(p["unidades_vendidas"]), float(p["ingresos"])
        ], thin_border)
    for col, ancho in zip("ABCDEF", [15, 35, 20, 12, 15, 15]):
        ws.column_dimensions[col].width = ancho


def _excel_hoja_inventario(ws, datos, header_font, header_fill, center_align, thin_border):
    """Rellena la hoja Excel para el reporte de inventario."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.title = "Inventario Actual"
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"Reporte de Inventario — {datos['generado_en'].strftime('%d/%m/%Y %H:%M')}"
    c.font      = Font(bold=True, size=14, color="1A2942")
    c.alignment = center_align
    c.fill      = PatternFill("solid", fgColor="EDF1F7")

    r = datos["resumen"]
    for row_num, (label, value) in enumerate([
        ("Total Productos Activos", int(r["total_productos"])),
        ("Total Unidades en Stock", int(r["total_unidades"])),
        ("Valor Total Inventario",  float(r["valor_inventario"])),
        ("Productos Críticos",      int(r["productos_criticos"])),
    ], start=3):
        ws.cell(row=row_num, column=1, value=label)
        ws.cell(row=row_num, column=2, value=value)

    ws.cell(row=8, column=1, value="DETALLE").font = Font(bold=True, size=11)
    _excel_escribir_header(ws, 9,
        [TITULO_CODIGO, "Producto", "Categoría", "Precio", "Stock Actual", "Stock Mínimo", "Estado"],
        header_font, header_fill, center_align, thin_border)
    for i, p in enumerate(datos["detalle"], start=10):
        _excel_escribir_datos(ws, i, [
            p["codigo"], p["nombre"], p["categoria"],
            float(p["precio_unitario"]), p["stock_actual"],
            p["stock_minimo"], p["estado_stock"]
        ], thin_border)
    for col, ancho in zip("ABCDEFG", [14, 35, 18, 12, 13, 13, 12]):
        ws.column_dimensions[col].width = ancho


def _excel_hoja_vendedor(wb, ws, datos, header_font, header_fill, center_align, thin_border):
    """Rellena la hoja Excel para el reporte por vendedor."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.title = "Ventas por Vendedor"
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"Reporte por Vendedor — {datos['periodo']['inicio']} al {datos['periodo']['fin']}"
    c.font      = Font(bold=True, size=14, color="1A2942")
    c.alignment = center_align
    c.fill      = PatternFill("solid", fgColor="EDF1F7")

    ws.cell(row=3, column=1, value="RENDIMIENTO").font = Font(bold=True, size=11)
    _excel_escribir_header(ws, 4,
        ["Vendedor", TITULO_VENTAS, TITULO_INGRESOS, TITULO_TICKET, "Venta Máxima"],
        header_font, header_fill, center_align, thin_border)
    for i, v in enumerate(datos["por_vendedor"], start=5):
        _excel_escribir_datos(ws, i, [
            v["vendedor_nombre"], int(v["total_ventas"]),
            float(v["ingresos_totales"]),
            round(float(v["ticket_promedio"]), 2),
            float(v["venta_maxima"])
        ], thin_border)
    for col, ancho in zip("ABCDE", [30, 14, 16, 16, 14]):
        ws.column_dimensions[col].width = ancho

    ws_det = wb.create_sheet("Detalle de Ventas")
    _excel_escribir_header(ws_det, 1,
        ["Factura", "Fecha", "Vendedor", "Cliente", "Total", "Método Pago"],
        header_font, header_fill, center_align, thin_border)
    for i, d in enumerate(datos["detalle"], start=2):
        _excel_escribir_datos(ws_det, i, [
            d["numero_factura"],
            d["fecha_hora"].strftime("%d/%m/%Y %H:%M"),
            d["vendedor_nombre"], d["cliente_nombre"],
            float(d["total"]), d["metodo_pago"]
        ], thin_border)
    for col, ancho in zip("ABCDEF", [20, 18, 28, 28, 14, 16]):
        ws_det.column_dimensions[col].width = ancho


def exportar_excel(tipo: str, datos: dict, ruta: str) -> tuple[bool, str]:
    """Exporta un reporte a formato .xlsx usando openpyxl."""
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        hf, hfill, ca, tb = _excel_estilos(openpyxl)

        if tipo == "ventas":
            _excel_hoja_ventas(ws, datos, hf, hfill, ca, tb)
        elif tipo == "inventario":
            _excel_hoja_inventario(ws, datos, hf, hfill, ca, tb)
        elif tipo == "vendedor":
            _excel_hoja_vendedor(wb, ws, datos, hf, hfill, ca, tb)

        wb.save(ruta)
        return True, f"Excel guardado en {ruta}"
    except ImportError:
        return False, "openpyxl no está instalado. Ejecuta: pip install openpyxl"
    except Exception as e:
        return False, f"Error al exportar Excel: {e}"


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN A PDF
# ══════════════════════════════════════════════════════════════════════════════

def _pdf_estilos(colors, styles):
    """Retorna los estilos de párrafo para el PDF."""
    from reportlab.lib.styles import ParagraphStyle
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"], fontSize=18,
        textColor=colors.HexColor("#1A2942"), spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionHeader", parent=styles["Heading2"], fontSize=13,
        textColor=colors.HexColor("#FF6B2B"), spaceBefore=14, spaceAfter=6,
    )
    return title_style, section_style


def _pdf_tabla(colors, encabezados, filas, col_widths=None):
    """Helper que crea una tabla con estilo corporativo para PDF."""
    from reportlab.platypus import Table, TableStyle
    data = [encabezados] + filas
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1A2942")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  10),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFD")]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E6F0")),
        ("FONTSIZE",      (0, 1), (-1, -1), 9),
        ("ROWHEIGHT",     (0, 0), (-1, -1), 20),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def _pdf_seccion_ventas(story, datos, colors, styles, title_style, section_style, inch):
    """Genera los elementos PDF para el reporte de ventas."""
    from reportlab.platypus import Paragraph, Spacer
    story.append(Paragraph(TITULO_AUTOPARTS, title_style))
    story.append(Paragraph(
        f"Reporte de Ventas | {datos['periodo']['inicio']} — {datos['periodo']['fin']}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.2 * inch))

    r = datos["resumen"]
    story.append(Paragraph("Resumen del Período", section_style))
    filas_res = [
        [TITULO_VENTAS,         str(int(r["total_ventas"]))],
        [TITULO_INGRESOS,       f"${float(r['ingresos_totales']):,.2f}"],
        ["Descuentos Aplicados", f"${float(r['descuentos_totales']):,.2f}"],
        [TITULO_TICKET,         f"${float(r['ticket_promedio']):,.2f}"],
    ]
    story.append(_pdf_tabla(colors, ["Métrica", "Valor"], filas_res,
                            col_widths=[3 * inch, 3 * inch]))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("Top 10 Productos Más Vendidos", section_style))
    filas_top = [
        [p["codigo"], p["producto_nombre"][:28],
         str(int(p["unidades_vendidas"])), f"${float(p['ingresos']):,.2f}"]
        for p in datos["productos_top"]
    ]
    story.append(_pdf_tabla(colors, [TITULO_CODIGO, "Producto", "Unidades", "Ingresos"],
                            filas_top,
                            col_widths=[1.2*inch, 3.2*inch, 1.2*inch, 1.4*inch]))


def _pdf_seccion_inventario(story, datos, colors, styles, title_style, section_style, inch):
    """Genera los elementos PDF para el reporte de inventario."""
    from reportlab.platypus import Paragraph, Spacer
    story.append(Paragraph(TITULO_AUTOPARTS, title_style))
    story.append(Paragraph(
        f"Reporte de Inventario | {datos['generado_en'].strftime('%d/%m/%Y %H:%M')}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.2 * inch))
    r = datos["resumen"]
    story.append(Paragraph("Resumen de Inventario", section_style))
    filas_res = [
        ["Productos Activos",    str(int(r["total_productos"]))],
        ["Total Unidades",       str(int(r["total_unidades"]))],
        ["Valor del Inventario", f"${float(r['valor_inventario']):,.2f}"],
        ["Productos Críticos",   str(int(r["productos_criticos"]))],
    ]
    story.append(_pdf_tabla(colors, ["Métrica", "Valor"], filas_res,
                            col_widths=[3 * inch, 3 * inch]))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("Productos con Stock Crítico o Sin Stock", section_style))
    if datos["criticos"]:
        filas_crit = [
            [p["codigo"], p["nombre"][:28], p["categoria"],
             str(p["stock_actual"]), str(p["stock_minimo"]), p["estado_stock"]]
            for p in datos["criticos"]
        ]
        story.append(_pdf_tabla(colors,
            [TITULO_CODIGO, "Producto", "Cat.", "Stock", "Mín.", "Estado"], filas_crit,
            col_widths=[1.1*inch, 2.6*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1.0*inch]))
    else:
        story.append(Paragraph("No hay productos en estado crítico.", styles["Normal"]))


def _pdf_seccion_vendedor(story, datos, colors, styles, title_style, section_style, inch):
    """Genera los elementos PDF para el reporte por vendedor."""
    from reportlab.platypus import Paragraph, Spacer
    story.append(Paragraph(TITULO_AUTOPARTS, title_style))
    story.append(Paragraph(
        f"Reporte por Vendedor | {datos['periodo']['inicio']} — {datos['periodo']['fin']}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("Rendimiento por Vendedor", section_style))
    filas_v = [
        [v["vendedor_nombre"], str(int(v["total_ventas"])),
         f"${float(v['ingresos_totales']):,.2f}",
         f"${float(v['ticket_promedio']):,.2f}"]
        for v in datos["por_vendedor"]
    ]
    story.append(_pdf_tabla(colors,
        ["Vendedor", "Ventas", "Ingresos", "Ticket Prom."], filas_v,
        col_widths=[2.8*inch, 1.2*inch, 1.8*inch, 1.6*inch]))


def exportar_pdf(tipo: str, datos: dict, ruta: str) -> tuple[bool, str]:
    """Exporta un reporte a PDF usando reportlab."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate
        from reportlab.lib.styles import getSampleStyleSheet

        doc    = SimpleDocTemplate(ruta, pagesize=letter)
        styles = getSampleStyleSheet()
        story  = []
        title_style, section_style = _pdf_estilos(colors, styles)

        secciones = {
            "ventas":     _pdf_seccion_ventas,
            "inventario": _pdf_seccion_inventario,
            "vendedor":   _pdf_seccion_vendedor,
        }
        if tipo in secciones:
            secciones[tipo](story, datos, colors, styles, title_style, section_style, inch)

        doc.build(story)
        return True, f"PDF guardado en {ruta}"
    except ImportError:
        return False, "reportlab no está instalado. Ejecuta: pip install reportlab"
    except Exception as e:
        return False, f"Error al exportar PDF: {e}"
